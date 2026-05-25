from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from travel_planner.db import DEFAULT_DB_PATH, init_database, get_connection

MONTHS = [
    "Januar",
    "Februar",
    "Marec",
    "April",
    "Maj",
    "Junij",
    "Julij",
    "Avgust",
    "September",
    "Oktober",
    "November",
    "December",
]

COUNTRY_ROWS = range(4, 30)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().lower().replace(",", ".")
        if value in {"", "x", "/"}:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def upsert_country(conn, name: str) -> int:
    conn.execute("INSERT OR IGNORE INTO countries (name) VALUES (?)", (name,))
    row = conn.execute("SELECT id FROM countries WHERE name = ?", (name,)).fetchone()
    return int(row["id"])


def insert_travel_option(
    conn,
    country_id: int,
    mode: str,
    label: str,
    month: int | None,
    cost: float | None,
    price_scope: str,
    source_note: str,
) -> None:
    if cost is None:
        return
    conn.execute(
        """
        INSERT INTO travel_options
            (country_id, mode, label, month, cost_eur, price_scope, source_note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (country_id, mode, label, month, cost, price_scope, source_note),
    )


def import_travel_to_destination(conn, ws) -> None:
    for row in COUNTRY_ROWS:
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        country_id = upsert_country(conn, country)

        for index, month_name in enumerate(MONTHS, start=1):
            cost = to_number(ws.cell(row=row, column=2 + index).value)
            insert_travel_option(
                conn,
                country_id,
                "flight_ljubljana",
                "Letalo iz Ljubljane",
                index,
                cost,
                "person",
                f"Letalska karta iz Ljubljane, {month_name}",
            )

        road_country = clean_text(ws.cell(row=row, column=17).value)
        if road_country:
            road_country_id = upsert_country(conn, road_country)
            insert_travel_option(
                conn,
                road_country_id,
                "road",
                "Cestni prevoz",
                None,
                to_number(ws.cell(row=row, column=18).value),
                "group",
                "Strošek prevoza po podatkih ViaMichelin",
            )

    for row in range(38, 64):
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        country_id = upsert_country(conn, country)
        for index, month_name in enumerate(MONTHS, start=1):
            insert_travel_option(
                conn,
                country_id,
                "flight_zagreb",
                "Letalo iz Zagreba",
                index,
                to_number(ws.cell(row=row, column=2 + index).value),
                "person",
                f"Letalska karta iz Zagreba, {month_name}",
            )

    for row in range(39, 65):
        country = clean_text(ws.cell(row=row, column=17).value)
        if not country:
            continue
        insert_travel_option(
            conn,
            upsert_country(conn, country),
            "train",
            "Vlak",
            None,
            to_number(ws.cell(row=row, column=18).value),
            "person",
            "Cena karte za vlak",
        )

    for row in range(74, 100):
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        insert_travel_option(
            conn,
            upsert_country(conn, country),
            "bus",
            "Avtobus",
            None,
            to_number(ws.cell(row=row, column=3).value),
            "person",
            "Cena karte za avtobus",
        )


def import_lodging(conn, ws) -> None:
    for row in COUNTRY_ROWS:
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        country_id = upsert_country(conn, country)
        for persons in range(1, 6):
            cost = to_number(ws.cell(row=row, column=2 + persons).value)
            if cost is None:
                continue
            conn.execute(
                """
                INSERT INTO lodging_rates (country_id, persons, nightly_cost_eur)
                VALUES (?, ?, ?)
                """,
                (country_id, persons, cost),
            )


def import_local_transport(conn, ws) -> None:
    for row in COUNTRY_ROWS:
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        conn.execute(
            """
            INSERT OR REPLACE INTO local_transport_costs (
                country_id,
                bus_daily_eur,
                tram_daily_eur,
                metro_daily_eur,
                train_daily_eur,
                taxi_10km_eur,
                rental_car_daily_eur,
                bike_daily_eur,
                walkability_score
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                upsert_country(conn, country),
                to_number(ws.cell(row=row, column=3).value),
                to_number(ws.cell(row=row, column=4).value),
                to_number(ws.cell(row=row, column=5).value),
                to_number(ws.cell(row=row, column=6).value),
                to_number(ws.cell(row=row, column=7).value),
                to_number(ws.cell(row=row, column=8).value),
                to_number(ws.cell(row=row, column=9).value),
                to_number(ws.cell(row=row, column=10).value),
            ),
        )


def import_activities(conn, ws) -> None:
    for row in range(3, 29):
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        conn.execute(
            """
            INSERT OR REPLACE INTO activity_costs
                (country_id, culture_daily_eur, nature_daily_eur)
            VALUES (?, ?, ?)
            """,
            (
                upsert_country(conn, country),
                to_number(ws.cell(row=row, column=3).value) or 0,
                to_number(ws.cell(row=row, column=4).value) or 0,
            ),
        )


def import_food(conn, ws) -> None:
    for row in range(5, 31):
        country = clean_text(ws.cell(row=row, column=2).value)
        if not country:
            continue
        conn.execute(
            """
            INSERT OR REPLACE INTO food_costs
                (country_id, restaurant_daily_eur, fastfood_daily_eur, grocery_daily_eur)
            VALUES (?, ?, ?, ?)
            """,
            (
                upsert_country(conn, country),
                to_number(ws.cell(row=row, column=3).value) or 0,
                to_number(ws.cell(row=row, column=4).value) or 0,
                to_number(ws.cell(row=row, column=5).value) or 0,
            ),
        )


def read_existing_users(db_path: Path) -> list[dict[str, Any]]:
    if not db_path.exists():
        return []
    try:
        with get_connection(db_path) as conn:
            rows = conn.execute(
                "SELECT id, name, email, password_hash, created_at FROM users"
            ).fetchall()
            return [dict(row) for row in rows]
    except Exception:
        return []


def restore_existing_users(conn, users: list[dict[str, Any]]) -> None:
    for user in users:
        conn.execute(
            """
            INSERT OR IGNORE INTO users (id, name, email, password_hash, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                user["id"],
                user["name"],
                user["email"],
                user["password_hash"],
                user["created_at"],
            ),
        )


def import_workbook(workbook_path: Path, db_path: Path = DEFAULT_DB_PATH) -> dict[str, int]:
    existing_users = read_existing_users(Path(db_path))
    init_database(db_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)

    with get_connection(db_path) as conn:
        restore_existing_users(conn, existing_users)
        import_travel_to_destination(conn, workbook["Pot na destinacijo"])
        import_lodging(conn, workbook["Prenočišča"])
        import_local_transport(conn, workbook["Prevoz na destinaciji"])
        import_activities(conn, workbook["Aktivnosti"])
        import_food(conn, workbook["Prehrana"])

        conn.commit()
        return {
            "countries": conn.execute("SELECT COUNT(*) AS count FROM countries").fetchone()["count"],
            "travel_options": conn.execute("SELECT COUNT(*) AS count FROM travel_options").fetchone()["count"],
            "lodging_rates": conn.execute("SELECT COUNT(*) AS count FROM lodging_rates").fetchone()["count"],
            "local_transport": conn.execute(
                "SELECT COUNT(*) AS count FROM local_transport_costs"
            ).fetchone()["count"],
            "activities": conn.execute("SELECT COUNT(*) AS count FROM activity_costs").fetchone()["count"],
            "food": conn.execute("SELECT COUNT(*) AS count FROM food_costs").fetchone()["count"],
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import travel Excel data into SQLite.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default=Path("data/raw/travel_data.xlsx"),
        type=Path,
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        type=Path,
        help="Path to the SQLite database.",
    )
    args = parser.parse_args()
    counts = import_workbook(args.workbook, args.db)
    print(counts)


if __name__ == "__main__":
    main()
